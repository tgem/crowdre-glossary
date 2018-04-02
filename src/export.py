# -*- coding: utf-8 -*-
"""
Created on Sat Dec  2 09:39:48 2017

@author: Tim G
"""

"""
Export
"""

import openpyxl as pyxl

def print_index(index):
    for term in index:
        print(term)

def save_index_as_text(index):
    with open('..\\target\\terms.txt', 'w') as f:
        for term in index:
            termstring = ""
            for element in term:
                termstring = termstring + element + " "
            f.write(termstring+':'+str(len(index[term]))+"\n")

def save_index_as_xlsx(index, reqs, ids):
    wb = pyxl.Workbook()
    ws = wb.active
    ws.title = 'Requirements'
    ws['A1']='ID'
    ws['B1']='Requirement'
    for i, req in enumerate(reqs):
        ws['A'+str(i+2)]=int(ids[i])
        ws['B'+str(i+2)]=req
    ws = wb.create_sheet('Glossary',0)
    ws['A1']='Glossary term'
    ws['B1']='Number of related requirements'
    i=2
    for term in index:
        termstring = ""
        for element in term:
            termstring = termstring + element + " "
        ws['A'+str(i)]=termstring
        ws['B'+str(i)]=len(index[term])
        i=i+1
    ws = wb.create_sheet('Index',0)
    ws['A1']='Glossary term'
    ws['B1']='Requirement'
    i=2
    for term in index:
        termstring = ""
        for element in term:
            termstring = termstring + element + " "
        for id in index[term]:
            ws['A'+str(i)]=termstring
            ws['B'+str(i)]='=VLOOKUP('+str(ids[id])+',Requirements!A2:B2967,2,FALSE)'
            i=i+1    
    ws.auto_filter.ref = "A1:B"+str(i)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    wb.save('..\\target\\glossary.xlsx')
